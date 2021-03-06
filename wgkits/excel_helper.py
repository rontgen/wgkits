# -*- coding:utf-8 -*
"""
author:rontgen(wanggeng)
Date: 2018-08-08
Function Description:
    1. read, write, add, del, modify, query operation for excel files
    2. read data from excel, analyze and draw graphic chart
"""

import openpyxl as ox

import os, sys, string
from wgkits.build_tool import *
import math
from wgkits.overload import overload

ref=string.ascii_uppercase

cur_dir = os.getcwd()

def read_xlsx(file_path):
    if sys.version_info <(3,0):
        file_path = unicode(path2unix(file_path), 'utf8')
    wb = ox.load_workbook(file_path)
    return wb

def write_xlsx(file_path, wb):
    if sys.version_info <(3,0):
        file_path = unicode(path2unix(file_path), 'utf8')
    folder_name = os.path.dirname(file_path)
    if not has_dir(folder_name):
        os.makedirs(folder_name)
    wb.save(file_path)

@overload(ox.worksheet.worksheet.Worksheet, tuple)
def read_cell(ws, tp):
    return ws.cell(tp[0], tp[1]).value

@overload(ox.worksheet.worksheet.Worksheet, int, int)
def read_cell(ws, r, c):
    return ws.cell(row=r, column=c).value

@overload(ox.worksheet.worksheet.Worksheet, str)
def read_cell(ws, cor):
    p=re.compile(r'^([A-Za-z]+)(\d+)$')
    m = p.match(cor)
    if m:
         return read_cell(ws, cor2rc(cor))
    else:
        raise TypeError("params not match")

@overload(ox.worksheet.worksheet.Worksheet,tuple, str)
def write_cell(ws, tp, data):
    ws.cell(tp[0], tp[1]).value = data

@overload(ox.worksheet.worksheet.Worksheet, int, int ,str)
def write_cell(ws, r, c, data):
    ws.cell(row=r, column=c).value = data

@overload(ox.worksheet.worksheet.Worksheet, str, str)
def write_cell(ws, cor, data):
    p=re.compile(r'^([A-Za-z]+)(\d+)$')
    m = p.match(cor)
    if m:
         write_cell(ws, cor2rc(cor), data)
    else:
        raise TypeError("params not match")


def insert_row(ws, row_index, data_list):
    row = ws.max_row
    column = ws.max_column
    cur_index = row
    if len(data_list) != column:
        sys_quit("params not match")
    if row_index <= row:
        row_list = []
        if sys.version_info < (3, 0):
            row_list = map(lambda x : x+1, range(row))
        else:
            row_list = list(map(lambda x : x+1, range(row)))
        for index in row_list[::-1]:
            if index < row_index:
                break
            else:
                col_index = 1
                for c in ws.iter_cols():
                    data = ""
                    if sys.version_info < (3, 0):
                        data = str(read_cell(ws, index, col_index)).encode('utf8')
                    else:
                        data = str(read_cell(ws, index, col_index))
                    if read_cell(ws, index, col_index) != None:
                        write_cell(ws,index + 1, col_index, data)
                    else:
                        write_cell(ws,index + 1, col_index, "")
                    col_index = col_index + 1

    col_index = 1
    for c in ws.iter_cols():
        write_cell(ws, row_index, col_index, data_list[col_index-1])
        col_index = col_index + 1

def insert_column(ws, col_index, data_list):
    row = ws.max_row
    column = ws.max_column
    if len(data_list) > row:
        sys_quit("params not match")
    elif len(data_list) < row:
        data_list = data_list + [""] * (row-len(data_list))
    if col_index <= column:
        column_list = []
        if sys.version_info < (3, 0):
            column_list = map(lambda x : x+1, range(column))
        else:
            column_list = list(map(lambda x : x+1, range(column)))
        for index in column_list[::-1]:
            if index < col_index:
                break
            else:
                row_index = 1
                for c in ws.iter_rows():
                    data = ""
                    if sys.version_info < (3, 0) and read_cell(ws, row_index, index) != None :
                        data = read_cell(ws, row_index, index).strip().encode('utf-8')
                    else:
                        data = str(read_cell(ws, row_index, index))
                    if read_cell(ws, row_index, index) != None:
                        write_cell(ws, row_index, index+1, data)
                    else:
                        write_cell(ws, row_index, index+1, "")
                    row_index = row_index + 1
    row_index = 1
    for c in ws.iter_rows():
        write_cell(ws, row_index, col_index, data_list[row_index - 1])
        row_index = row_index + 1

def modify_xlsx():
    pass

def query_xlsx_coordinate():
    pass

def query_xlsx_data():
    pass

def rc2cor(row, column):
    remainder = 0
    ret = 0
    rl = []
    result=""
    pow_num = 10
    lenlist=0
    while True:
        rt = column - math.pow(26, pow_num)
        if rt > 0:
            break
        else:
            pow_num = pow_num -1
    print("pow_num=%s" %pow_num)
    lenlist = pow_num
    while True:
        if pow_num < 0:
            break
        print("[start]:remainder=%d" %remainder)
        print("[start]:ret=%d" %ret)
        print("[start]:column=%d" %column)
        print("[start]:pow_num=%d" %pow_num)
        ret = int(math.floor(column / math.pow(26, pow_num)))
        remainder = column % 26
        if ret > 0:
            print("remainder=%d" %remainder)
            print("ret=%d" %ret)
            print("column=%d" %column)
            print("pow_num=%d" %pow_num)
            if remainder != 0:
                rl.append(ret)
                column = column - math.pow(26, pow_num)
                remainder = 0
                pow_num = pow_num - 1
            else:
                print("special # rl={0}" .format(rl))
                rl.append(ret)
                print("@@@ ret={0}" .format(ret))
                print("@@@ rl={0}" .format(len(rl)))
                if ret > len(rl):
                    polish_list = [0] * (ret-len(rl))
                    remainder=""
                    print("@@@ polish_list={0}" .format(polish_list))
                    rl = rl + polish_list
                    print("@@special={0}" .format(rl))
                    break
        else:
            if remainder > 0:
                rl.append(ret)
            ret = ""
            print("remainder=%s" %remainder)
            if remainder >= 1:
                remainder = ref[int(round(remainder)) -1]
            else:
                remainder=""
            break

    print("caculate rl=%s" %rl)
    print("lenlist=%s" %lenlist)
    index=-1
    if len(rl) < lenlist:
        tmp = [0] * (lenlist-len(rl))
        rl=rl+tmp
        print("###rl={0}" .format(rl))
    for m in rl[::-1]:
        index=index-1
        print("index=%d" %index)
        if -index > len(rl):
            break
        if m == 0 and rl[index]>0:
            rl[index+1]=rl[index+1]+26
            rl[index]=rl[index]-1
        elif m==0 and rl[index]==0:
            rl[index] = -1
            rl[index+1] = rl[index+1]+26
        print("reorder list=%s" %rl)
    for m in rl[::-1]:
        if m >0:
            result=ref[m-1]+result
    print("result=%s" %type(result))
    print("remainder=%s" %type(remainder))
    print("row=%s" %type(row))
    return result+remainder+ str(row)

def letter2num(lstr):
    rt = 0
    digit = 1
    for m in lstr[::-1]:
        rt = rt +(ref.index(m)+1)*math.pow(26, digit-1)
        digit = digit + 1
    return int(round(rt))

def cor2rc(corstr):
    rt = (-1,-1)
    p = re.compile(r'(?P<clm>\D*)(?P<row>\d*)')
    m = p.match(corstr)
    if m:
        print(m.groupdict()['clm'])
        print(m.groupdict()['row'])
        rt = (int(m.groupdict()['row']), letter2num(m.groupdict()['clm']))
    return rt

def main():
    file_path = path2unix(os.path.join(cur_dir, "../新建文本文档.xlsx"))
    wb = read_xlsx(file_path)
    sheet_list = wb.sheetnames
    print(sheet_list)
    sh1=sheet_list[0]
    sheet = wb.get_sheet_by_name(sh1)
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.coordinate)
            print(cell.value)

    print(sheet.max_row, sheet.max_column)
    print(wb.active)
    print(rc2cor(3,677))
    #print(math.pow(26, 10))


if __name__ == '__main__':
    main()
